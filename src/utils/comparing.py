# utils/comparing.py

import os
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Dict, List
import src.config as config
import math
import re
from src.Tomato import Tomato
from src.config import dictionaries
from src.config.enums import ImagePathKind, ModelType
from datetime import datetime

def verify_lab_results(tomatoes: List[Tomato], output_excel_path: str = 'project_data/Data/comparing/Lab_Results_Comparison.xlsx') -> None:
    """
    Compares the Lab_Results.csv file with the provided Tomato dataset and creates
    a copy of the CSV with discrepancies marked in red in an Excel file.

    Args:
        tomatoes (List): List of Tomato objects.
        output_excel_path (str, optional): Path to save the comparison Excel file.
                                           Defaults to 'Lab_Results_Comparison.xlsx'.
    """
    # Define the mapping between CSV columns and Tomato attributes
    column_mapping = {
        'Weight (g)': 'weight',
        'Firmness': 'firmness',
        'Citric Acid (%)': 'citric_acid',
        'pH': 'pH',
        'TSS': 'TSS',
        'Ascrobic acid (mg/100g)': 'ascorbic_acid',
        'area(cm^2)': 'spectral_stats.area'
    }

    # Define the fill for discrepancies (red)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Read the Lab_Results.csv file
    try:
        lab_results_df = pd.read_csv(config.LAB_RESULT_FILENAME)
    except Exception as e:
        print(f"Error reading {config.LAB_RESULT_FILENAME}: {e}")
        return

    # Read the CSV into Excel
    excel_path = output_excel_path
    lab_results_df.to_excel(excel_path, index=False)
    print(f"Copied Lab_Results.csv to {excel_path} for comparison.")

    # Load the Excel file with openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Determine if the dataset is single-sided or two-sided
    images_path_kind = config.IMAGES_PATH_KIND.value
    if images_path_kind not in [ImagePathKind.ONE_SIDE.value, ImagePathKind.TWO_SIDES.value]:
        print("Error: config.IMAGES_PATH_KIND must be either 'oneSide' or 'twoSides'.")
        return

    # Create a mapping from _id to Tomato object for quick access
    id_to_tomato = {tomato._id: tomato for tomato in tomatoes}

    # Iterate over the DataFrame rows
    for index, row in lab_results_df.iterrows():
        internal_id = row.get('internal_id')
        if pd.isna(internal_id):
            print(f"Row {index + 2} has missing internal_id. Skipping.")
            continue

        # Map internal_id to _id based on images_path_kind
        if images_path_kind == 'oneSide':
            # For single-sided images, _id = ceil(internal_id / 2)
            try:
                _id = math.ceil(int(internal_id) / 2)
            except Exception as e:
                print(f"Error processing internal_id '{internal_id}' at row {index + 2}: {e}")
                continue
        else:
            # For two-sided images, _id = internal_id
            try:
                _id = int(internal_id)
            except Exception as e:
                print(f"Error processing internal_id '{internal_id}' at row {index + 2}: {e}")
                continue

        # Get the corresponding Tomato object
        tomato = id_to_tomato.get(_id)
        if not tomato:
            print(f"No Tomato object found with _id {_id} for internal_id {internal_id} at row {index + 2}.")
            # Optionally, mark the entire row red or skip
            continue

        # Get the quality_assess and spectral_stats.area from the Tomato object
        quality_assess = tomato.quality_assess
        spectral_stats = tomato.spectral_stats

        # Iterate through the relevant columns to compare
        for col_idx, csv_column in enumerate(lab_results_df.columns, 1):
            if csv_column not in column_mapping:
                continue  # Skip columns that are not being compared

            dataset_attr = column_mapping[csv_column]

            # Get the value from the Tomato object
            if dataset_attr == 'spectral_stats.area':
                expected_value = getattr(spectral_stats, 'area', None)
            else:
                expected_value = getattr(quality_assess, dataset_attr, None)

            # Get the value from the CSV
            csv_value = row[csv_column]

            # Check for missing values in the CSV
            if pd.isna(csv_value):
                cell = ws.cell(row=index + 2, column=col_idx)
                cell.fill = red_fill
                continue

            # Round floating numbers to three decimal places
            if isinstance(csv_value, float):
                csv_value_rounded = round(csv_value, 3)
            elif isinstance(csv_value, int):
                csv_value_rounded = float(csv_value)
            else:
                # Attempt to convert to float if possible
                try:
                    csv_value_rounded = round(float(csv_value), 3)
                except:
                    # If conversion fails, skip comparison
                    continue

            if dataset_attr == 'spectral_stats.area':
                if isinstance(expected_value, float):
                    expected_value_rounded = round(expected_value, 3)
                elif isinstance(expected_value, int):
                    expected_value_rounded = float(expected_value)
                else:
                    # Attempt to convert to float if possible
                    try:
                        expected_value_rounded = round(float(expected_value), 3)
                    except:
                        expected_value_rounded = None

                # Compare the values
                if expected_value_rounded != csv_value_rounded:
                    cell = ws.cell(row=index + 2, column=col_idx)
                    cell.fill = red_fill

            else:
                # Compare other quality_assess attributes
                if isinstance(expected_value, float):
                    expected_value_rounded = round(expected_value, 3)
                elif isinstance(expected_value, int):
                    expected_value_rounded = float(expected_value)
                else:
                    # Attempt to convert to float if possible
                    try:
                        expected_value_rounded = round(float(expected_value), 3)
                    except:
                        expected_value_rounded = None

                # Compare the values
                if expected_value_rounded != csv_value_rounded:
                    cell = ws.cell(row=index + 2, column=col_idx)
                    cell.fill = red_fill

        # Optionally, compare other columns or handle 'ID' differently

    # Save the modified Excel file
    wb.save(output_excel_path)
    print(f"Comparison completed. The output file with discrepancies is saved as '{output_excel_path}'.")


def compare_model_metrics_to_best_results(model_metrics: Dict[str, Dict[str, float]], selected_bands: List[int], model_name: str, model_type: ModelType, output_filename: str = None) -> None:
    """
    Compares the model's test metrics to precomputed XGBoost metrics from metrics_dict.

    Args:
        model_metrics (Dict[str, Dict[str, float]]): Metrics obtained from model.test(), structured as {attribute: {metric: value}}.
        selected_bands (List[int]): List of selected band indices used in the model.
    """
    # Convert selected_bands to a sorted tuple for consistent key matching
    selected_bands_tuple = tuple(sorted(selected_bands))

    # Check if the selected band combination exists in metrics_dict
    if config.COMPARE_TO_BEST_RESULTS:
        xgb_metrics = dictionaries.metrics_dict.get('best_result')
        if not xgb_metrics:
            print("No best_result metrics found in metrics_dict.")
    else:
        if selected_bands_tuple in dictionaries.metrics_dict:
            xgb_metrics = dictionaries.metrics_dict[selected_bands_tuple].get(config.ModelType.XGBOOST.value)
        else:
            print(f"No metrics found in metrics_dict for bands: {selected_bands_tuple}, default bands will be selected")
            xgb_metrics = dictionaries.metrics_dict[(104, 113, 170)].get(config.ModelType.XGBOOST.value)
        

    # Define the mapping between model_metrics keys and metrics_dict keys
    metric_mapping = {'R²': 'r^2', 'RMSE': 'rmse'}

    # Extract the list of quality attributes (exclude 'average' as it's a computed metric)
    attributes = [attr for attr in model_metrics.keys() if attr != 'average']

    # Define which metrics to compare
    metrics_to_compare = config.COMPARING_MATRICS
    num_metrics = len(metrics_to_compare)

    # Create subplots for each metric
    fig, axes = plt.subplots(1, num_metrics, figsize=(6 * num_metrics, 6))

    # Ensure axes is iterable
    if num_metrics == 1:
        axes = [axes]

    for ax, metric_name_model, metric_name_xgb in zip(axes, metrics_to_compare, [metric_mapping[m] for m in metrics_to_compare]):
        # Extract model and XGBoost values for the current metric
        model_values = [model_metrics[attr].get(metric_name_model, np.nan) for attr in attributes]
        xgb_values = [xgb_metrics[attr].get(metric_name_xgb, np.nan) for attr in attributes]

        # Define the position of bars on the x-axis
        x = np.arange(len(attributes))
        width = 0.35  # Width of the bars

        # Create bars for the model and XGBoost
        rects1 = ax.bar(x - width/2, model_values, width, label=f'Our - {model_type.value}', color='skyblue')
        rects2 = ax.bar(x + width/2, xgb_values, width, label='Best Result', color='salmon')

        # Add labels, title, and custom x-axis tick labels
        ax.set_ylabel(metric_name_model)
        ax.set_title(f'{metric_name_model} - {model_name}')
        ax.set_xticks(x)
        ax.set_xticklabels(attributes, rotation=45, ha='right')
        ax.legend()

        # Function to attach a text label above each bar
        def autolabel(rects):
            for rect in rects:
                height = rect.get_height()
                if not np.isnan(height):
                    ax.annotate(f'{height:.2f}',
                                xy=(rect.get_x() + rect.get_width() / 2, height),
                                xytext=(0, 3),  # 3 points vertical offset
                                textcoords="offset points",
                                ha='center', va='bottom')

        # Attach labels to both sets of bars
        autolabel(rects1)
        autolabel(rects2)

    # Adjust layout to prevent clipping of tick-labels
    fig.tight_layout()

    # Check if output_filename is provided and save the plot
    if output_filename and output_filename.strip() and os.path.dirname(output_filename):
        save_dir = os.path.dirname(output_filename)
        os.makedirs(save_dir, exist_ok=True)  # Create the directory if it doesn't exist
        save_path = os.path.join(save_dir, f'{model_name}.png')
        fig.savefig(save_path)
        print(f'Figure saved to {save_path}')

    # Display the plot
    plt.show(block=config.IMAGES_BLOCKER)

def save_dl_results_to_csv(
    model_metrics: Dict[str, Dict[str, float]], 
    selected_bands: List[int], 
    model_name: str, 
    model_type: ModelType, 
    csv_path: str = None
) -> None:
    """
    Save DL model results to the CSV file with both R² and RMSE values.
    
    Args:
        model_metrics: Dictionary containing metrics for each quality attribute
        selected_bands: List of selected band indices
        model_name: Name of the model
        model_type: Type of the model
        csv_path: Path to the CSV file
    """
    # Use centralized path if csv_path not provided
    if csv_path is None:
        from src.config.paths import CSV_RESULTS_PATH
        csv_path = CSV_RESULTS_PATH
    
    # Create results directory if it doesn't exist
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    
    # Safe getter function with proper defaults
    def safe_get_config(attr, default):
        return getattr(config, attr, default)
    
    # Get indexes safely
    indexes = safe_get_config('INDEXES', [])
    has_indexes = isinstance(indexes, list) and len(indexes) > 0
    
    # Get NDSI pairs safely
    ndsi_pairs = safe_get_config('NDSI_BAND_PAIRS', [])
    ndsi_pairs_str = ';'.join([f'({pair[0]},{pair[1]})' for pair in ndsi_pairs]) if ndsi_pairs else ''
    
    # Get index names safely
    index_names_str = ''
    if has_indexes:
        index_names = []
        for idx in indexes:
            if hasattr(idx, 'name'):
                index_names.append(str(idx.name))
            else:
                index_names.append(str(idx))
        index_names_str = ','.join(index_names)
    
    # Prepare the row data based on CSV structure
    row_data = {
        'model_type': model_type.value,
        'regression_mode': 'multi' if safe_get_config('MULTI_REGRESSION', True) else 'single',
        'reflectance': str(safe_get_config('REFLECTANCE', True)).upper(),
        'std': str(safe_get_config('STD', True)).upper(),
        'ndsi': str(safe_get_config('NDSI', True)).upper(),
        'indexes': str(has_indexes).upper(),
        'model_depth': safe_get_config('MODEL_DEPTH', 18),
        'reflectance_bands': ','.join(map(str, selected_bands)),
        'std_bands': ','.join(map(str, selected_bands)),
        'ndsi_band_pairs': ndsi_pairs_str,
        'index_names': index_names_str,
        'augment_times': safe_get_config('AUGMENT_TIMES', 0),
        'padded_value': safe_get_config('PADDED_VALUE', 0),
        'patience': safe_get_config('PATIENCE', 7),
        'use_validation': str(safe_get_config('USE_VALIDATION', True)).upper(),
        'date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'empty_1': '',
        'empty_2': '',
    }
    
    # Add quality attribute metrics (R² values)
    quality_attributes = ['TSS', 'citric_acid', 'firmness', 'pH', 'weight', 'ascorbic_acid']
    
    for attr in quality_attributes:
        if attr in model_metrics:
            r2_value = model_metrics[attr].get('R²', 0.0)
            rrmse_value = model_metrics[attr].get('RRMSE', 0.0)  # Fixed: Use 'RRMSE' not 'RMSE'
            row_data[attr] = round(float(r2_value), 4)
            row_data[f'{attr}_RMSE'] = round(float(rrmse_value), 4)  # Column name stays _RMSE for compatibility
        else:
            row_data[attr] = 0.0
            row_data[f'{attr}_RMSE'] = 0.0
    
    # Calculate average score (still based on R² only)
    valid_r2_values = [float(row_data[attr]) for attr in quality_attributes if float(row_data[attr]) > 0]
    avg_score = sum(valid_r2_values) / len(valid_r2_values) if valid_r2_values else 0.0
    row_data['score'] = round(avg_score, 2)
    
    # Read existing CSV or create new one
    try:
        df = pd.read_csv(csv_path)
    except FileNotFoundError:
        # Create new DataFrame with headers including RMSE columns
        columns = [
            'model_type', 'regression_mode', 'reflectance', 'std', 'ndsi', 'indexes', 
            'model_depth', 'reflectance_bands', 'std_bands', 'ndsi_band_pairs', 
            'index_names', 'augment_times', 'padded_value', 'patience', 'use_validation', 
            'date', 'empty_1', 'empty_2', 
            # R² columns
            'TSS', 'citric_acid', 'firmness', 'pH', 'weight', 'ascorbic_acid', 
            'empty_3',  # Separator before RMSE columns
            # RMSE columns
            'TSS_RMSE', 'citric_acid_RMSE', 'firmness_RMSE', 'pH_RMSE', 'weight_RMSE', 'ascorbic_acid_RMSE',
            'score'
        ]
        df = pd.DataFrame(columns=columns)
        print(f"📄 Created new CSV file with R² and RMSE columns: {csv_path}")
    
    # Add empty columns if they don't exist (for backward compatibility)
    if 'empty_3' not in df.columns.tolist():
        # Insert empty_3 after ascorbic_acid
        try:
            ascorbic_pos = df.columns.tolist().index('ascorbic_acid')
            insert_pos = ascorbic_pos + 1
        except ValueError:
            insert_pos = len(df.columns)
        
        df.insert(insert_pos, 'empty_3', '')
        
        # Add RMSE columns after empty_3
        for i, attr in enumerate(quality_attributes):
            rmse_col = f'{attr}_RMSE'
            if rmse_col not in df.columns.tolist():
                df.insert(insert_pos + 1 + i, rmse_col, 0.0)
        
        print(f"📊 Added RMSE columns to existing CSV file")
    
    # Fill missing RMSE columns with 0 for existing rows
    for attr in quality_attributes:
        rmse_col = f'{attr}_RMSE'
        if rmse_col in df.columns.tolist():
            df[rmse_col] = df[rmse_col].fillna(0.0)
    
    # Add the new row
    new_row = pd.DataFrame([row_data])
    df = pd.concat([df, new_row], ignore_index=True)
    
    # Save back to CSV
    df.to_csv(csv_path, index=False)
    print(f"✅ Results saved to {csv_path}")
    print(f"📊 Model: {model_name}, Average R²: {avg_score:.4f}")
    
    # Print rRMSE summary
    rrmse_values = [row_data[f'{attr}_RMSE'] for attr in quality_attributes if row_data[f'{attr}_RMSE'] > 0]
    if rrmse_values:
        avg_rrmse = sum(rrmse_values) / len(rrmse_values)
        print(f"📊 Average rRMSE: {avg_rrmse:.4f}")


def export_tomato_reflectance_band_to_excel(tomato, band_index, output_dir="./exports"):
    """
    Extracts a specific band from tomato reflectance data and exports it to Excel.
    
    Args:
        tomato (Tomato): A Tomato object containing spectral_stats
        band_index (int): The specific band index to extract
        output_dir (str): Directory where the Excel file will be saved
    
    Returns:
        str: Path to the created Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create a filename with tomato info
    filename = f"tomato_id{tomato._id}_band{band_index}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Create Excel writer with pandas
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Check if reflectance_matrix is a dictionary (two sides) or array (one side)
        reflectance_matrix = tomato.spectral_stats.reflectance_matrix
        
        if isinstance(reflectance_matrix, dict):
            # Two-sided tomato
            for side in ['sideA', 'sideB']:
                if reflectance_matrix.get(side) is not None:
                    # Extract the 2D matrix for the specified band
                    band_matrix = reflectance_matrix[side][:, :, band_index]
                    
                    # Convert to DataFrame and save as a sheet
                    df = pd.DataFrame(band_matrix)
                    df.to_excel(writer, sheet_name=side, index=False)
                    
        else:
            # Single-sided tomato (just in case, though from your description we expect two sides)
            band_matrix = reflectance_matrix[:, :, band_index]
            df = pd.DataFrame(band_matrix)
            df.to_excel(writer, sheet_name='reflectance', index=False)
    
    print(f"Exported reflectance data for tomato ID {tomato._id}, band {band_index} to {filepath}")
    return filepath
